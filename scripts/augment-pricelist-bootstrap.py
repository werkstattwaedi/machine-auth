#!/usr/bin/env python3
# Copyright Offene Werkstatt Wädenswil
# SPDX-License-Identifier: MIT
"""
Augment Mario's pricelist workbook with the columns our catalog importer needs,
WITHOUT touching his calculation formulas.

Mario maintains one workbook with a sheet per workshop. He curates two label
columns per row — `Etikett Name` and `Etikett Mass` — that we import verbatim
(they drive the printed label). This script *appends* four columns the importer
reads by header name:

    Code            stable 4-digit article number (catalog identity)
    Kategorie       top-level category  (from the section headings)
    Unterkategorie  sub-category        (from sub-headings / species)
    Einheit         sale unit (m² / lm / kg / Stk) → pricing model

The sale PRICE is not copied — the importer reads Mario's existing
"Preis Einheit Verkauf" column by header name (single source of truth). The
display `name` is composed by the importer from `Etikett Name` + `Etikett Mass`,
so no `Name` column is injected here.

It also NORMALISES the curated labels for consistency (trailing spaces, unit
spacing "24mm"→"24 mm", dimensions "15x2mm"→"15 × 2 mm") and writes the cleaned
values back into the Etikett cells so the file Mike shares back to Mario is
consistent. Every change is listed in the reconciliation report.

Workflow: run this → Mike shares the produced "… – mit Codes.xlsx" back to
Mario → from then on the injected columns travel with his file. Re-running is
idempotent: an existing `Code` value on a row is preserved; only new rows get a
fresh number.

It also generates two sheets Mario's file lacks: a `Makerspace` product sheet
(one-time migration from makerspace.json — literal prices, no formulas) and a
global `Varianten` sheet that defines the laser cut-to-size factors. From here on
`Varianten` is the editorial source for those factors (they used to be hardcoded
in shared/src/pricing.ts); the importer expands each item's variants from it.

Output: a sibling "… – mit Codes.xlsx" plus a reconciliation report.
"""
import json
import re
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

SRC = Path("/mnt/c/Users/Mike/Downloads/260616 Preisberechnungsliste.xlsx")
OUT = SRC.with_name(SRC.stem + " – mit Codes.xlsx")
REPORT = Path("/tmp/pricelist_reconciliation.md")
SEED = Path(__file__).resolve().parent / "seed-data" / "catalog"
OPS = Path(__file__).resolve().parent.parent.parent / "machine-auth-operations" / "scripts" / "seed-data" / "catalog"

# ── section → pricing-model (human sale-unit label) ──────────────────────────
# Holz mixes units by section; the others are uniform per sheet.
HOLZ_TOPS = [
    "Massivholz", "Holzplatten", "Dübel- und Rundstäbe",
    "Schleifmittel", "Holzverbinder und Kleinteile", "Varia",
]
HOLZ_BANNERS = ["Massivholz und Holzplatten"]
HOLZ_EINHEIT = {
    "Massivholz": "m²", "Holzplatten": "m²", "Dübel- und Rundstäbe": "lm",
    "Schleifmittel": "Stk", "Holzverbinder und Kleinteile": "Stk", "Varia": "Stk",
}

# kind: "holz" = section tree + species-as-subcategory for Massivholz;
#       "flat" = each section heading is a top category, no sub-category.
SHEETS = {
    "Holz":    {"kind": "holz", "seed": "holz.json",    "fresh_base": 3156},
    "Metall":  {"kind": "flat", "seed": None,           "fresh_base": 2001, "einheit": "lm"},
    "Keramik": {"kind": "flat", "seed": "keramik.json", "fresh_base": 4216, "einheit": "kg"},
    "Textil":  {"kind": "flat", "seed": None,           "fresh_base": 7001, "einheit": "Stk"},
    "Glas":    {"kind": "flat", "seed": "glas.json",    "fresh_base": 5503, "einheit": "Stk"},
    "Stein":   {"kind": "flat", "seed": None,           "fresh_base": 8001, "einheit": "Stk"},
    "Schmuck": {"kind": "flat", "seed": None,           "fresh_base": 8501, "einheit": "Stk"},
}

NOTE_MARKERS = ("Preisliste", "Verkaufspreis =", "Marge", "Bestellliste",
                "noch offen", "Einkaufspreis")

# ── Makerspace + variant definitions (generated once from the seed) ──────────
MAKERSPACE_JSON = "makerspace.json"

# Variant definitions seeded once into the workbook's `Varianten` sheet. From
# here on that sheet is the editorial source the importer reads (these values
# previously lived hardcoded in shared/src/pricing.ts). id, label, factor,
# base pricing model of the derived variant. Factors are the cut area in m²;
# a3 carries a small handling margin over the exact 0.12474 area (matches the
# ad-hoc laser pricelist). 100×100 / 500×1000 are the acryl cut formats.
VARIANT_DEFS = [
    ("a3", "Zuschnitt A3", 0.126, "count"),
    ("320-620", "Zuschnitt 320 × 620 mm", 0.1984, "count"),
    ("500-1250", "Zuschnitt 500 × 1250 mm", 0.625, "count"),
    ("100-100", "Zuschnitt 100 × 100 mm", 0.01, "count"),
    ("500-1000", "Zuschnitt 500 × 1000 mm", 0.5, "count"),
]

# base pricingModel → human sale unit (Einheit column, importer contract).
PRICINGMODEL_TO_EINHEIT = {
    "area": "m²", "weight": "kg", "sla": "L", "length": "lm", "count": "Stk",
}

report = []


def norm(v):
    return "" if v is None else str(v).replace("\n", " ").strip()


def key(s):
    """Match key: collapse whitespace, lowercase."""
    return re.sub(r"\s+", " ", str(s).strip()).lower()


def norm_mass(raw):
    """Consistent mass spelling: '24mm'→'24 mm', '15x2mm'→'15 × 2 mm'."""
    s = re.sub(r"\s+", " ", norm(raw))
    if not s:
        return s
    dim = re.match(
        r"^(\d+(?:\.\d+)?)\s*[x×X]\s*(\d+(?:\.\d+)?)"
        r"(?:\s*[x×X]\s*(\d+(?:\.\d+)?))?\s*(mm|cm|m)?$",
        s,
    )
    if dim:
        nums = [g for g in dim.groups()[:3] if g]
        return " × ".join(nums) + " " + (dim.group(4) or "mm")
    single = re.match(r"^(\d+(?:\.\d+)?)\s*(mm|cm|m|g|kg|ml|l)$", s, re.I)
    if single:
        return f"{single.group(1)} {single.group(2).lower()}"
    return s


def norm_name(raw):
    return re.sub(r"\s+", " ", norm(raw))


def split_name_mass(name):
    """Split a seed `name` into (Etikett Name, Etikett Mass) on a trailing
    `<n>mm` token: 'Sperrholz Pappel 12mm' → ('Sperrholz Pappel', '12 mm');
    'Filament Standard' → ('Filament Standard', '')."""
    n = norm_name(name)
    m = re.match(r"^(.*?)\s+(\d+(?:\.\d+)?\s*mm)$", n)
    if m:
        return norm_name(m.group(1)), norm_mass(m.group(2))
    return n, ""


def load_seedmap(fname):
    """norm(name) → code from the prod (ops) seed, falling back to the public one."""
    if not fname:
        return {}
    path = OPS / fname if (OPS / fname).exists() else SEED / fname
    if not path.exists():
        return {}
    return {key(i["name"]): str(i["code"]) for i in json.load(open(path))}


def header_row(ws):
    for r in range(1, 20):
        if any("Preis Einheit Verkauf" in norm(ws.cell(r, c).value)
               for c in range(1, ws.max_column + 1)):
            return r
    return None


def col_of(ws, hr, label):
    for c in range(1, ws.max_column + 1):
        if label in norm(ws.cell(hr, c).value):
            return c
    return None


def is_note(a):
    return any(m in a for m in NOTE_MARKERS)


def parse_sheet(ws, cfg):
    """Yield product dicts {row, name, mass, kategorie, unterkategorie, einheit}."""
    hr = header_row(ws)
    if hr is None:
        return [], None, {}
    cols = {
        "hr": hr,
        "produkt": col_of(ws, hr, "Produkt"),
        "name": col_of(ws, hr, "Etikett Name"),
        "mass": col_of(ws, hr, "Etikett Mass"),
        "code": col_of(ws, hr, "Code"),  # None on first run
    }
    products = []
    top = sub = ""
    # Scan from row 1 so a section title sitting *above* the header row (e.g.
    # "Tone"/"Siebdruck"/"Blankstahl") still sets the category; only rows below
    # the header are considered products.
    for r in range(1, ws.max_row + 1):
        a = norm(ws.cell(r, 1).value)
        name = norm(ws.cell(r, cols["name"]).value) if cols["name"] else ""
        produkt = norm(ws.cell(r, cols["produkt"]).value) if cols["produkt"] else ""
        # Skip Mario's repeated header blocks mid-sheet.
        if name == "Etikett Name" or produkt == "Produkt":
            continue
        # A product row is any row below the header carrying a real Etikett Name.
        if r > hr and name and name != "#N/A":
            if cfg["kind"] == "holz":
                kategorie = top
                unterkategorie = produkt if top == "Massivholz" else sub
                einheit = HOLZ_EINHEIT.get(top, "Stk")
            else:
                kategorie = top
                unterkategorie = ""
                einheit = cfg["einheit"]
            products.append({
                "row": r,
                "name": norm_name(name),
                "mass": norm_mass(ws.cell(r, cols["mass"]).value) if cols["mass"] else "",
                "kategorie": kategorie or "Sonstiges",
                "unterkategorie": unterkategorie,
                "einheit": einheit,
            })
            continue
        # Otherwise a heading iff col A carries non-note text.
        if a and not is_note(a) and a not in HOLZ_BANNERS:
            if cfg["kind"] == "holz":
                if a in HOLZ_TOPS:
                    top, sub = a, ""
                else:
                    sub = a
            else:
                top, sub = a, ""
    return products, cols, {}


def assign_codes(sheet, products, cfg):
    seedmap = load_seedmap(cfg["seed"])
    used = set()
    nxt = cfg["fresh_base"]
    matched = new = 0
    for p in products:
        k = key(f"{p['name']} {p['mass']}".strip())
        code = seedmap.get(k)
        if code and code not in used:
            matched += 1
        else:
            code = str(nxt)
            nxt += 1
            new += 1
        p["code"] = code
        used.add(code)
    return matched, new


def last_content_col(ws, hr):
    last = 1
    for c in range(1, ws.max_column + 1):
        if norm(ws.cell(hr, c).value):
            last = c
    return last


def makerspace_seed_path():
    """makerspace.json from the prod (ops) seed, falling back to the public one."""
    return OPS / MAKERSPACE_JSON if (OPS / MAKERSPACE_JSON).exists() else SEED / MAKERSPACE_JSON


def build_varianten_sheet(wb):
    """Create the global `Varianten` definition sheet the importer reads:
    id | label | factor | base pricing model of the derived variant. Guarded —
    once the sheet exists it's the editorial source (Mario/Mike maintain the
    factors there), so re-runs leave it untouched."""
    if "Varianten" in wb.sheetnames:
        return None
    ws = wb.create_sheet("Varianten")
    for c, h in enumerate(["Variante", "Bezeichnung", "Faktor", "Grundmodell"], start=1):
        ws.cell(1, c, h)
    for r, (vid, label, factor, model) in enumerate(VARIANT_DEFS, start=2):
        ws.cell(r, 1, vid)
        ws.cell(r, 2, label)
        ws.cell(r, 3, factor)
        ws.cell(r, 4, model)
    return len(VARIANT_DEFS)


def build_makerspace_sheet(wb):
    """One-time seed→sheet migration: build a flat `Makerspace` product sheet
    from makerspace.json with the full importer contract. Prices are literal
    values (no formula), so this sheet needs no Excel recalculation. Skipped
    once Mario's workbook already carries a Makerspace sheet (he owns it then)."""
    if "Makerspace" in wb.sheetnames:
        report.append("**Makerspace**: Blatt bereits vorhanden — unverändert gelassen.\n")
        return 0
    path = makerspace_seed_path()
    if not path.exists():
        report.append("**Makerspace**: keine `makerspace.json` gefunden — übersprungen.\n")
        return 0
    items = sorted(json.load(open(path)), key=lambda i: int(i["code"]))
    ws = wb.create_sheet("Makerspace")
    headers = ["Code", "Kategorie", "Unterkategorie", "Einheit",
               "Etikett Name", "Etikett Mass", "Preis Einheit Verkauf", "Varianten"]
    for c, h in enumerate(headers, start=1):
        ws.cell(1, c, h)
    variant_rows = 0
    for r, item in enumerate(items, start=2):
        cat = item.get("category", [])
        variants = item["variants"]
        base = variants[0]
        ename, emass = split_name_mass(item["name"])
        # Cut options are the non-base variants; their ids feed the Varianten
        # column (the seed stores fully-expanded variants, not variantIds).
        vids = [v["id"] for v in variants[1:]]
        if vids:
            variant_rows += 1
        ws.cell(r, 1, int(item["code"]))  # numeric — see note in the augment loop
        ws.cell(r, 2, cat[0] if cat else "Sonstiges")
        ws.cell(r, 3, cat[1] if len(cat) > 1 else "")
        ws.cell(r, 4, PRICINGMODEL_TO_EINHEIT.get(base["pricingModel"], "Stk"))
        ws.cell(r, 5, ename)
        ws.cell(r, 6, emass)
        ws.cell(r, 7, base["unitPrice"]["default"])
        ws.cell(r, 8, ",".join(vids))
    codes = sorted(int(i["code"]) for i in items)
    report.append(
        f"**Makerspace**: {len(items)} Produkte aus `{path.name}` erzeugt. "
        f"Codes {codes[0]}–{codes[-1]}. {variant_rows} Zeilen mit Varianten "
        f"(Preise als Festwerte, keine Formeln).\n"
    )
    return len(items)


def main():
    wb_vals = openpyxl.load_workbook(SRC, data_only=True)
    wb = openpyxl.load_workbook(SRC, data_only=False)

    total = 0
    for sheet, cfg in SHEETS.items():
        products, cols, _ = parse_sheet(wb_vals[sheet], cfg)
        if not products:
            report.append(f"**{sheet}**: keine Produkte mit `Etikett Name` — übersprungen.")
            continue
        matched, new = assign_codes(sheet, products, cfg)
        total += len(products)

        ws = wb[sheet]
        hr = cols["hr"]
        anchor = last_content_col(wb_vals[sheet], hr) + 2  # blank gap column
        headers = ["Code", "Kategorie", "Unterkategorie", "Einheit"]
        for i, h in enumerate(headers):
            ws.cell(hr, anchor + i, h)

        label_edits = 0
        cat_samples = {}
        for p in products:
            r = p["row"]
            # Write cleaned labels back into Mario's Etikett cells.
            if cols["name"]:
                if norm(wb_vals[sheet].cell(r, cols["name"]).value) != p["name"]:
                    label_edits += 1
                ws.cell(r, cols["name"], p["name"])
            if cols["mass"]:
                if norm(wb_vals[sheet].cell(r, cols["mass"]).value) != p["mass"]:
                    label_edits += 1
                ws.cell(r, cols["mass"], p["mass"])
            # Numeric cell (codes are 4-digit, no leading zero) so Excel doesn't
            # flag "number stored as text"; the importer coerces to string anyway.
            ws.cell(r, anchor + 0, int(p["code"]))
            ws.cell(r, anchor + 1, p["kategorie"])
            ws.cell(r, anchor + 2, p["unterkategorie"])
            ws.cell(r, anchor + 3, p["einheit"])
            cat = " / ".join(x for x in (p["kategorie"], p["unterkategorie"]) if x)
            cat_samples.setdefault(cat, p["code"])

        letter = get_column_letter(anchor)
        codes = sorted(int(p["code"]) for p in products)
        report.append(
            f"**{sheet}**: {len(products)} Produkte — {matched} mit bestehendem Code, "
            f"{new} neu. Codes {codes[0]}–{codes[-1]}. "
            f"Spalten ab **{letter}**. {label_edits} Etikett-Zellen normalisiert."
        )
        report.append("  Kategorien: " + ", ".join(sorted(cat_samples)) + "\n")

    # Makerspace + variant definitions are generated (not augmented from Mario's
    # sheets): the Varianten sheet is the editorial source for variant factors,
    # the Makerspace sheet is a one-time seed→sheet migration.
    total += build_makerspace_sheet(wb)
    nvar = build_varianten_sheet(wb)
    if nvar is None:
        report.append("**Varianten**: Blatt bereits vorhanden — unverändert gelassen.\n")
    else:
        report.append(
            f"**Varianten**: {nvar} Definitionen (Referenzblatt, das der Import liest — "
            f"Faktoren leben ab jetzt hier, nicht mehr im Code).\n"
        )

    wb.save(OUT)

    md = [
        "# Preislisten-Bootstrap — Reconciliation\n",
        f"Quelle: `{SRC.name}`  →  Ausgabe: `{OUT.name}`\n",
        f"**{total} importierbare Zeilen.**\n",
        "## Pro Werkstatt\n",
        *[f"- {line}" for line in report],
        "\n## Hinweise\n",
        "- Codes sind ab jetzt stabil — nicht umnummerieren/wiederverwenden.\n",
        "- Glas/Stein/Schmuck: derzeit keine gültigen Produktzeilen (leer bzw. `#N/A`).\n",
        "- Preise stehen in Marios `Preis Einheit Verkauf` — vor dem Import in Excel "
        "öffnen und speichern, damit die Formeln berechnet sind.\n",
        "- Etikett Name/Mass wurden vereinheitlicht (bitte im File gegenlesen).\n",
        "- `Makerspace` + `Varianten` sind neu generiert: Makerspace-Preise sind "
        "Festwerte (kein Excel-Neuberechnen nötig); `Varianten` definiert die "
        "Zuschnitt-Faktoren, die der Import auf jede Zeile mit `Varianten`-Spalte "
        "anwendet.\n",
    ]
    REPORT.write_text("".join(md))
    print("".join(md))
    print(f"\n✓ geschrieben: {OUT}")


if __name__ == "__main__":
    main()
